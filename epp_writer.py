import openpyxl as xl
import epp_model as m

""" Static function to write and Excel containing the resutls of a peer evaluation (epp) """

def write_xlsx(filename: str, epp: m.EPP) -> None:
    """ write the content of the epp structure to an Excel file designated by filename.
        overwrite any existing file of the same name. """
    
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sommaire de l'EPP"
    
    # write header
    ws.append(build_header())
    
    # write data
    row_counter = 2
    for team in epp:
        first_row = row_counter
        rows = build_rows(team, row_counter)
        row_counter += len(team)
        for row in rows:
            ws.append(row)
        ws.merge_cells(start_row = first_row, start_column = 8, end_row = row_counter-1, end_column = 8)
        
    # set style
    for i in range(2, row_counter+1):
        c = ws.cell(i, 5)
        c.number_format = '0.00'
        c = ws.cell(i, 6)
        c.number_format = '0.00'
        c = ws.cell(i, 7)
        c.number_format = '0.00'
        c = ws.cell(i, 9)
        c.number_format = '0.00'
            
    # save to file
    wb.save(filename)
    
    
def build_rows(team: m.Team, row_counter: int) -> list:
    """ build rows containing en entire team """
    rows = []
    for index, e in enumerate(team):
        formula = get_formula(row_counter + index, row_counter)
        row = [team.name, e.last_name, e.surname, e.email, e.note, team.average, e.factor, "", formula]
        rows.append(row)

    return rows

def build_header() -> list:
    return ["Groupe", "Nom", "Prenom", "Courriel", "Note_EPP", "MNG", "Facteur", "Note_equipe", "Note_etudiant"]


def get_formula(first_row: int, row_counter: int):
    return "=G" + str(first_row) + "*H" + str(row_counter)
